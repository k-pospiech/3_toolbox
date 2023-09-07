from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def auto_adjust_columns(worksheet):
    """
    Auto-adjust the width of columns based on the content in each column.
    
    :param worksheet: The openpyxl worksheet object you want to adjust.
    """
    for col in worksheet.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def auto_adjust_row_heights(worksheet):
    """
    Auto-adjust the height of rows based on the content in each cell of the row.
    
    :param worksheet: The openpyxl worksheet object you want to adjust.
    """
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                num_lines = len(str(cell.value).split('\n'))
                worksheet.row_dimensions[cell.row].height = 15 * num_lines

def wrap_text_for_wide_columns(worksheet, max_width=25):
    """
    Enable text wrapping for columns wider than a given width.
    
    :param worksheet: The openpyxl worksheet object you want to adjust.
    :param max_width: Columns wider than this value will have text wrapping enabled. Default is 25.
    """
    for col_letter, dimension in worksheet.column_dimensions.items():
        if dimension.width > max_width:
            for cell in worksheet[col_letter]:
                cell.alignment = Alignment(wrap_text=True)

def center_text_in_all_cells(worksheet):
    """
    Center the text in all cells of the worksheet.
    
    :param worksheet: The openpyxl worksheet object you want to adjust.
    """
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def format_excel(file_path, sheet_name):
    """
    Apply multiple formatting options to an Excel sheet specified by file path and sheet name.
    
    :param file_path: Path to the Excel file you want to format.
    :param sheet_name: Name of the sheet within the Excel file you want to format.

    Example usage:
    >>> file_path = "your_file.xlsx"
    >>> sheet_name = "your_sheet_name"
    >>> format_excel(file_path, sheet_name)
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    auto_adjust_columns(ws)
    auto_adjust_row_heights(ws)
    wrap_text_for_wide_columns(ws)
    center_text_in_all_cells(ws)

    wb.save(file_path)


